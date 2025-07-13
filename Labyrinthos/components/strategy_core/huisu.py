#D:\MyMazeGame\TEST\4_password_test
import hashlib
import json
import os
from typing import List, Tuple, Dict, Set, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

# 固定的盐值
SALT = b'\xb2S"e}\xdf\xb0\xfe\x9c\xde\xde\xfe\xf3\x1d\xdc>'

def hash_password(password: str) -> str:
    """计算密码的SHA-256哈希值（带盐）"""
    password_bytes = password.encode('utf-8')
    hash_obj = hashlib.sha256(SALT + password_bytes)
    return hash_obj.hexdigest()

class PasswordSolver:
    def __init__(self, C: List[List[int]], L: str):
        self.C = C
        self.L = L
        self.sets = [set(range(10)), set(range(10)), set(range(10))]
        self.prime_constraint = False
        self.apply_clues()
    
    def apply_clues(self):
        """应用所有线索到候选数字集合"""
        for clue in self.C:
            if clue == [-1, -1]:
                self.prime_constraint = True
            elif len(clue) == 2:
                self.apply_parity_constraint(clue)
            elif len(clue) == 3:
                self.apply_fixed_constraint(clue)
        
        if self.prime_constraint:
            primes = {2, 3, 5, 7}
            for i in range(3):
                self.sets[i] = self.sets[i] & primes
    
    def apply_parity_constraint(self, clue: List[int]):
        """应用奇偶性约束"""
        pos, parity = clue
        idx = pos - 1
        if parity == 0:
            self.sets[idx] = self.sets[idx] & {0, 2, 4, 6, 8}
        elif parity == 1:
            self.sets[idx] = self.sets[idx] & {1, 3, 5, 7, 9}
    
    def apply_fixed_constraint(self, clue: List[int]):
        """应用固定数字约束"""
        for i, val in enumerate(clue):
            if val != -1:
                self.sets[i] = {val}
    
    def solve_method1(self) -> Tuple[str, int]:
        """方法1：标准顺序"""
        return self.backtrack_solve(method_id=1)
    
    def solve_method2(self) -> Tuple[str, int]:
        """方法2：交换5和6"""
        return self.backtrack_solve(method_id=2)
    
    def solve_method3(self) -> Tuple[str, int]:
        """方法3：逆序"""
        return self.backtrack_solve(method_id=3)
    
    def solve_method4(self) -> Tuple[str, int]:
        """方法4：逆位置顺序"""
        return self.backtrack_solve(method_id=4)
    
    def solve_method5(self) -> Tuple[str, int]:
        """方法5：中间优先"""
        return self.backtrack_solve(method_id=5)
    
    # def solve_method6(self) -> Tuple[str, int]:
    #     """方法6：智能频率排序"""
    #     digit_freq = {
    #         0: 0.05, 1: 0.15, 2: 0.15, 3: 0.15, 4: 0.15,
    #         5: 0.10, 6: 0.10, 7: 0.10, 8: 0.05, 9: 0.10
    #     }
        
    #     orders = [
    #         self.get_frequency_order(self.sets[0], digit_freq),
    #         self.get_frequency_order(self.sets[1], digit_freq),
    #         self.get_frequency_order(self.sets[2], digit_freq)
    #     ]
        
    #     tries = 0
    #     for d0 in orders[0]:
    #         for d1 in orders[1]:
    #             for d2 in orders[2]:
    #                 candidate = [d0, d1, d2]
    #                 if self.prime_constraint:
    #                     primes = {2, 3, 5, 7}
    #                     if not (set(candidate).issubset(primes) and len(set(candidate)) == 3):
    #                         continue
                    
    #                 password_str = ''.join(str(d) for d in candidate)
    #                 tries += 1
    #                 if hash_password(password_str) == self.L:
    #                     return password_str, tries
        
    #     raise ValueError("No valid password found")
    
    def solve_method6(self) -> Tuple[str, int]:
        """方法6：智能频率排序（基于实际密码数据优化）"""
        # 基于100个密码数据的实际频率分布
        position_freq = [
            # 百位频率分布
            {1: 0.09, 2: 0.09, 3: 0.11, 4: 0.05, 5: 0.12, 
             7: 0.19, 8: 0.10, 9: 0.09, 0: 0.07, 6: 0.04},
            # 十位频率分布
            {3: 0.19, 5: 0.16, 7: 0.16, 2: 0.09, 9: 0.08,
             0: 0.07, 1: 0.07, 4: 0.06, 8: 0.06, 6: 0.03},
            # 个位频率分布
            {3: 0.22, 5: 0.17, 7: 0.15, 2: 0.10, 9: 0.07,
             1: 0.06, 8: 0.06, 0: 0.05, 4: 0.04, 6: 0.03}
        ]
        
        # 为每个位置生成智能排序列表
        orders = []
        for i in range(3):
            candidates = list(self.sets[i])
            # 按该位置的实际频率降序排序
            candidates.sort(key=lambda d: -position_freq[i].get(d, 0.01))
            orders.append(candidates)
        
        tries = 0
        for d0 in orders[0]:
            for d1 in orders[1]:
                for d2 in orders[2]:
                    candidate = [d0, d1, d2]
                    # 检查素数约束（如果存在）
                    if self.prime_constraint:
                        primes = {2, 3, 5, 7}
                        if not (set(candidate).issubset(primes) and len(set(candidate))) == 3:
                            continue
                    
                    password_str = ''.join(str(d) for d in candidate)
                    tries += 1
                    if hash_password(password_str) == self.L:
                        return password_str, tries
        
        raise ValueError("No valid password found")
        
    def get_frequency_order(self, candidates: Set[int], freq: Dict[int, float]) -> List[int]:
        """根据数字频率生成智能排序"""
        candidate_freq = [(d, freq.get(d, 0.05)) for d in candidates]
        candidate_freq.sort(key=lambda x: x[1], reverse=True)
        return [d for d, _ in candidate_freq]
    
    def get_enum_order(self, method_id: int, candidates: Set[int]) -> List[int]:
        """根据方法ID生成枚举顺序"""
        if method_id == 1:
            return sorted(candidates)
        elif method_id == 2:
            lst = sorted(candidates)
            if 5 in lst and 6 in lst:
                idx5 = lst.index(5)
                idx6 = lst.index(6)
                lst[idx5], lst[idx6] = lst[idx6], lst[idx5]
            return lst
        elif method_id == 3:
            return sorted(candidates, reverse=True)
        elif method_id in (4, 5):
            return sorted(candidates)
    
    def get_position_order(self, method_id: int) -> List[int]:
        """根据方法ID获取位置顺序"""
        if method_id in (1, 2, 3):
            return [0, 1, 2]
        elif method_id == 4:
            return [2, 1, 0]
        elif method_id == 5:
            return [1, 0, 2]
    
    def backtrack_solve(self, method_id: int) -> Tuple[str, int]:
        """回溯法求解密码"""
        pos_order = self.get_position_order(method_id)
        
        orders = [
            self.get_enum_order(method_id, self.sets[pos_order[0]]),
            self.get_enum_order(method_id, self.sets[pos_order[1]]),
            self.get_enum_order(method_id, self.sets[pos_order[2]])
        ]
        
        tries = 0
        for d0 in orders[0]:
            for d1 in orders[1]:
                for d2 in orders[2]:
                    candidate = [0, 0, 0]
                    candidate[pos_order[0]] = d0
                    candidate[pos_order[1]] = d1
                    candidate[pos_order[2]] = d2
                    
                    if self.prime_constraint:
                        primes = {2, 3, 5, 7}
                        if not (set(candidate).issubset(primes) and len(set(candidate)) == 3):
                            continue
                    
                    password_str = ''.join(str(d) for d in candidate)
                    tries += 1
                    if hash_password(password_str) == self.L:
                        return password_str, tries
        
        raise ValueError("No valid password found")

def save_to_excel(results: List[Dict[str, Any]], output_file: str):
    """将结果保存到Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "密码解密结果"
    
    # 设置标题行
    headers = [
        "文件名", "密码", "方法1尝试次数", "方法2尝试次数", 
        "方法3尝试次数", "方法4尝试次数", "方法5尝试次数",
        "方法6尝试次数", "最小尝试次数", "验证哈希", "目标哈希"
    ]
    
    # 写入标题行并设置样式
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 写入数据行
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result["filename"])
        ws.cell(row=row, column=2, value=result["password"])
        ws.cell(row=row, column=3, value=result["results"]["method1"]["tries"])
        ws.cell(row=row, column=4, value=result["results"]["method2"]["tries"])
        # ws.cell(row=row, column=5, value=result["results"]["method3"]["tries"])
        ws.cell(row=row, column=6, value=result["results"]["method4"]["tries"])
        ws.cell(row=row, column=7, value=result["results"]["method5"]["tries"])
        ws.cell(row=row, column=8, value=result["results"]["method6"]["tries"])
        ws.cell(row=row, column=9, value=result["results"]["min_tries"])
        ws.cell(row=row, column=10, value=result["hash"])
        ws.cell(row=row, column=11, value=result["target_hash"])
    
    # 设置列宽
    column_widths = [15, 10, 15, 15, 15, 15, 15, 15, 15, 70, 70]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width
    
    # 保存文件
    wb.save(output_file)

def process_files(json_files: List[str], dir_path: str) -> Tuple[List[Dict[str, Any]], List[int]]:
    """处理所有JSON文件"""
    all_results = []
    all_min_tries = []
    
    for json_file in sorted(json_files):
        file_path = os.path.join(dir_path, json_file)
        try:
            with open(file_path, 'r') as f:
                data = json.load(f)
            
            C = data["C"]
            L = data["L"]
            expected_password = data.get("password", "")
            
            solver = PasswordSolver(C, L)
            
            password1, tries1 = solver.solve_method1()
            password2, tries2 = solver.solve_method2()
            password3, tries3 = solver.solve_method3()
            password4, tries4 = solver.solve_method4()
            password5, tries5 = solver.solve_method5()
            password6, tries6 = solver.solve_method6()
            
            assert password1 == password2 == password3 == password4 == password5 == password6
            password = password1
            
            if expected_password and password != expected_password:
                print(f"警告: {json_file} 解密密码不匹配! 预期: {expected_password}, 实际: {password}")
            
            min_tries = min(tries1, tries2, tries3, tries4, tries5, tries6)
            
            result = {
                "filename": json_file,
                "password": password,
                "results": {
                    "method1": {"tries": tries1, "password": [int(d) for d in password]},
                    "method2": {"tries": tries2, "password": [int(d) for d in password]},
                    # "method3": {"tries": tries3, "password": [int(d) for d in password]},
                    "method4": {"tries": tries4, "password": [int(d) for d in password]},
                    "method5": {"tries": tries5, "password": [int(d) for d in password]},
                    "method6": {"tries": tries6, "password": [int(d) for d in password]},
                    "min_tries": min_tries
                },
                "hash": hash_password(password),
                "target_hash": L
            }
            
            all_results.append(result)
            all_min_tries.append(min_tries)
            
            print(f"处理完成: {json_file}")
            
        except Exception as e:
            print(f"处理文件 {json_file} 时出错: {str(e)}")
    
    return all_results, all_min_tries

def main():
    """主函数"""
    print("密码解密程序 - 结果导出Excel")
    print("=" * 60)
    
    dir_path = input("请输入包含JSON文件的目录路径: ").strip()
    if not os.path.isdir(dir_path):
        print(f"错误: 目录 '{dir_path}' 不存在")
        return
    
    json_files = [
        f for f in os.listdir(dir_path) 
        if f.endswith('.json') and os.path.isfile(os.path.join(dir_path, f))
    ]
    
    if not json_files:
        print(f"目录 '{dir_path}' 中没有找到JSON文件")
        return
    
    print(f"找到 {len(json_files)} 个JSON文件，开始处理...")
    print("=" * 60)
    
    all_results, all_min_tries = process_files(json_files, dir_path)
    
    # 生成Excel文件名（带时间戳）
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"password_results_{timestamp}.xlsx"
    
    save_to_excel(all_results, output_file)
    
    # 输出统计信息
    if all_min_tries:
        avg_min_tries = sum(all_min_tries) / len(all_min_tries)
        min_of_mins = min(all_min_tries)
        max_of_mins = max(all_min_tries)
        
        print("\n统计结果:")
        print("=" * 60)
        print(f"处理文件总数: {len(all_min_tries)}")
        print(f"最小尝试次数范围: {min_of_mins} - {max_of_mins}")
        print(f"平均最小尝试次数: {avg_min_tries:.2f}")
        print(f"总最小尝试次数: {sum(all_min_tries)}")
        print(f"\n结果已保存到: {output_file}")
    
    print("\n所有文件处理完成！")

if __name__ == "__main__":
    main()

# D:\MyMazeGame\TEST\4_password_test